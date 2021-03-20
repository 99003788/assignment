import pandas as pd 
from openpyxl import load_workbook
import numpy as np 
import itertools
 
a1 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=0)
a2 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=1)
a3 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=2)
a4 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=3)
a5 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=4)
 
# a3 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=['Sheet3'])
# a4 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=['Sheet4'])
# a5 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=['Sheet5'])
 
# mdc=pd.concat(df_first_shift,axis=0,ignore_index=True)
# print(a1)
# print(a2)
# print(a3)
# print(a4)
# print(a5)
 
# merging the files
 
a6 = a1.merge(a2, on = "PS number", how = "left")
a7 = a6.merge(a3, on = "PS number", how = "left")
a8 = a7.merge(a4, on = "PS number", how = "left")
a9 = a8.merge(a5, on = "PS number", how = "left")
 
# print(a9)
 
#creating new excel file
# a9.to_excel("Output.xlsx", index = False)
# variable_1 = int(input("Enter PS number : "))
# print(variable_1)
variable_1 = int(input("Enter PS number : "))
 
pf_variable = pd.DataFrame(a9, columns= ['PS number','Display Name', 'Pin Code','Fone No.','Salary','Official Email Address'])
# print(pf_variable)
 
pf_variable.set_index('PS number', inplace=True)
 
res = pf_variable.loc[variable_1]
print(res)
# res.dropna()
 
path = r"D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx"
 
book  = load_workbook(path)
writer = pd.ExcelWriter(path, engine="openpyxl")
writer.book = book
 
if 'Mastersheet' in book.sheetnames:
    ref = book['Mastersheet']
    book.remove(ref)
 
res.to_excel(writer, sheet_name  = "Mastersheet")
 
writer.save()
writer.close()
# load_workbook.book = book





 
# read 2nd sheet of an excel file 
#dataframe2 = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=['Sheet1', 'Sheet2']) 
 
#print(dataframe2) 
 
#excel_file_1 = 'Book1_My_Example.xlsx'
#excel_file_2 = 'Book2_My_Example.xlsx'
 
# df_first = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book1_My_Example.xlsx', sheet_name=None)
# df_first.keys()
# odict_keys(['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5'])
# df_second_shift = pd.read_excel(r'D:\Python_Practice-1\15-03-2021\Read_Write_Excel\Book2_My_Example.xlsx')
