import openpyxl as xl


def extract():  # creating a method to accept input from the user
    input_list = []  # creating an empty list
    no_of_inputs = int(input("Enter number of inputs you want : "))  # asking for the total  umber of inputs

    for value in range(no_of_inputs):  # for storing inputs in the list
        input_value = input("Enter PS NUMBER:")  # asking for the input
        input_list.append(input_value)  # storing the inputs in the list
    return input_list  # returning the list containing all the inputs from the user


wb = xl.load_workbook("mini_project.xlsx")  # loading the excel file
sheet = wb.get_sheet_names()  # List names  of all sheets in the excel workbook
total_sheets = len(sheet)  # Calculating length of all sheets
list_of_sheets = []  # Creating an empty list for list of sheets

for total_sheet_index in range(total_sheets):
    list_of_sheets.append(wb[sheet[total_sheet_index]])  # adding all the sheets of the excel file in teh list

print(list_of_sheets)  # printing the list of sheets in the excel file
data_input = extract()  # invoking the extract method
new_row_index = 1
new_summary_index = 1
position = 0
count_column = 0
count_row = 0
master_sheet = wb['mastersheet']
summary_sheet = wb['summarysheet']
for mastersheet_row_index in range(1, master_sheet.max_row + 1):
    for mastersheet_column_index in range(1, master_sheet.max_column + 1):
        master_sheet.delete_rows(2, master_sheet.max_row)   # clearing all existing data present the mastersheet

for list_of_sheet_index in range(0, len(list_of_sheets)):
    if list_of_sheets[list_of_sheet_index] != wb['mastersheet'] and list_of_sheets[list_of_sheet_index] != wb['summarysheet']:
        sheet_row = list_of_sheets[list_of_sheet_index].max_row  # Calculating max rows
        sheet_column = list_of_sheets[list_of_sheet_index].max_column  # Calculating max columns
        for sheet_column_index in range(1, sheet_column + 1):
            master_sheet.cell(row=new_row_index, column=sheet_column_index).value = list_of_sheets[
                list_of_sheet_index].cell(row=1, column=sheet_column_index).value  # Copying all the headers from all the excel sheets into the mastersheet
        new_row_index += 1
        for data_input_index in range(len(data_input)):
            for sheet_row_index in range(1, sheet_row + 1):
                if list_of_sheets[list_of_sheet_index].cell(row=sheet_row_index, column=1).value == int(data_input[data_input_index]):  #
                    position = sheet_row_index  # Storing the position the input data value is present in the sheet
                    break
                else:
                    position = 0
            if position != 0:
                for sheet_column_index in range(1, sheet_column + 1):
                    master_sheet.cell(row=new_row_index, column=sheet_column_index).value = list_of_sheets[list_of_sheet_index].cell(row=position, column=sheet_column_index).value  # Copying values from each sheet to the mastersheet
                    count_column += 1
                count_row += 1
                new_row_index += 1
        new_row_index += 1
    else:
        continue

count_row = master_sheet.max_row    # Counting maximum number of rows in mastersheet
count_column = master_sheet.max_column - 1  # Counting maximum number of columns in mastersheet
summary_sheet.delete_rows(2, summary_sheet.max_row)
summary_sheet.cell(row=new_summary_index, column=1).value = 'No. of Rows'
summary_sheet.cell(row=new_summary_index, column=2).value = 'Column_number'
new_summary_index += 1
summary_sheet.cell(row=2, column=1).value = count_row  # Number of rows in the mastersheet
summary_sheet.cell(row=2, column=2).value = count_column    # Number of columns in the mastersheet

wb.save('mini_project.xlsx')
wb.close()

'''62173517
20718405
75500092 '''
